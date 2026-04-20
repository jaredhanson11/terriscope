"""Exports router."""

import io

import openpyxl
import openpyxl.utils
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select

from src.app.database import DatabaseSession
from src.models.graph import LayerModel, MapModel, NodeModel, ZipAssignmentModel
from src.services.auth import CurrentUserDependency
from src.services.permissions import PermissionsServiceDependency

exports_router = APIRouter(prefix="/maps", tags=["Exports"])


@exports_router.get("/{map_id}/export/ztt")
def export_ztt(
    map_id: str,
    db: DatabaseSession,
    current_user: CurrentUserDependency,
    permission_service: PermissionsServiceDependency,
) -> StreamingResponse:
    """Export a map's zip-to-territory hierarchy as an Excel (.xlsx) file.

    Columns: zip_code | <layer order=1 name> | <layer order=2 name> | …
    One row per assigned zip code, with the full ancestor chain filled in.
    """
    if not permission_service.check_for_map_access(
        user_id=current_user.id,
        map_id=map_id,
        map_roles=["OWNER"],
    ):
        raise HTTPException(403)

    map_model = db.get(MapModel, map_id)
    if not map_model:
        raise HTTPException(404, "Map not found.")

    layers = (
        db.execute(
            select(LayerModel)
            .where(LayerModel.map_id == map_id)
            .order_by(LayerModel.order)
        )
        .scalars()
        .all()
    )
    if not layers:
        raise HTTPException(404, "No layers found for this map.")

    zip_layer = next((la for la in layers if la.order == 0), None)
    if not zip_layer:
        raise HTTPException(404, "No zip layer found.")

    upper_layers = [la for la in layers if la.order >= 1]

    # Build a flat node lookup: node_id → (name, parent_node_id, layer_id)
    node_lookup: dict[int, tuple[str, int | None, int]] = {}
    if upper_layers:
        upper_layer_ids = [la.id for la in upper_layers]
        for row in db.execute(
            select(NodeModel.id, NodeModel.name, NodeModel.parent_node_id, NodeModel.layer_id)
            .where(NodeModel.layer_id.in_(upper_layer_ids))
        ).all():
            node_lookup[row.id] = (row.name, row.parent_node_id, row.layer_id)

    # Layer id → layer object (for ancestor column mapping)
    layer_by_id = {la.id: la for la in upper_layers}

    # All assigned zip codes for this map's zip layer
    zip_assignments = db.execute(
        select(ZipAssignmentModel.zip_code, ZipAssignmentModel.parent_node_id)
        .where(ZipAssignmentModel.layer_id == zip_layer.id)
        .where(ZipAssignmentModel.parent_node_id.isnot(None))
        .order_by(ZipAssignmentModel.zip_code)
    ).all()

    # Build rows by walking the ancestor chain for each zip
    ExportRow = dict[str, str]
    rows: list[ExportRow] = []
    for za in zip_assignments:
        row: ExportRow = {"zip_code": za.zip_code}
        for ul in upper_layers:
            row[ul.name] = ""

        current_id: int | None = za.parent_node_id
        while current_id is not None:
            if current_id not in node_lookup:
                break
            node_name, parent_id, layer_id = node_lookup[current_id]
            if layer_id in layer_by_id:
                row[layer_by_id[layer_id].name] = node_name
            current_id = parent_id

        rows.append(row)

    # ── Build Excel workbook ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = map_model.name[:31]  # Excel sheet name max = 31 chars

    headers = ["zip_code"] + [la.name for la in upper_layers]

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[row_idx].height = 16

    # Auto-fit column widths (capped at 40)
    for col_idx, header in enumerate(headers, start=1):
        col_max = max(
            len(header),
            max((len(row.get(header, "")) for row in rows), default=0),
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(col_max + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    slug = map_model.name.replace(" ", "_").lower()
    filename = f"{slug}_ztt.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
